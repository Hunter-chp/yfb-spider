# -*- coding: utf-8 -*-
import requests
import pandas as pd
import random
import pickle
import os
import sys
import re
import time
from datetime import datetime, timedelta
from urllib.parse import urlencode
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import ctypes
import ctypes.wintypes
import winsound

# ==================== 配置 Selenium Manager 国内镜像源 ====================
os.environ['SE_MANAGER_MIRROR'] = 'https://registry.npmmirror.com/-/binary'
os.environ['SE_MANAGER_DEBUG'] = 'false'  # 关闭 Selenium Manager 调试输出
print("已配置 Selenium Manager 使用国内镜像源。")
# ==================== 配置结束 ====================

# --- Selenium 相关导入 ---
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.safari.options import Options as SafariOptions
from selenium.common.exceptions import WebDriverException

# ==================== 全局配置 ====================
EXCEL_FILE_NAME = '乙方宝采购公告汇总.xlsx'
OUTPUT_DIR = 'output'
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

COOKIE_FILE = 'session_cookies.pkl'

BASE_API_URL = "https://qiye.qianlima.com/new_qd_yfbsite/api/search"
DETAIL_API_COMMON = "https://qiye.qianlima.com/new_qd_yfbsite/api/subZhaobiao/zbDetail"

global_driver = None

# ==================== 自定义重启异常 ====================
class RestartException(Exception):
    """用于触发程序重启的异常"""
    pass

def input_with_restart(prompt):
    """带重启指令的输入函数，如果输入 @restart 则抛出异常"""
    value = input(prompt)
    if value.strip() == '@restart':
        raise RestartException()
    return value

# ==================== 地区数据（根据用户提供）====================
PROVINCE_ID_MAP = {
    "全国": "2703", "北京": "2", "天津": "26", "上海": "24", "重庆": "31",
    "香港": "3557", "澳门": "3558", "台湾": "3559", "广东": "5", "河北": "9",
    "山西": "22", "内蒙古": "18", "辽宁": "17", "吉林": "14", "黑龙江": "11",
    "江苏": "15", "浙江": "30", "安徽": "1", "福建": "3", "江西": "16",
    "山东": "21", "河南": "10", "湖北": "12", "湖南": "13", "广西": "6",
    "海南": "8", "四川": "25", "贵州": "7", "云南": "29", "西藏": "27",
    "陕西": "23", "甘肃": "4", "青海": "20", "宁夏": "19", "新疆": "28",
}

CITY_ID_MAP = {
    # 广东
    "广州": "305", "潮州": "302", "东莞": "303", "佛山": "304", "河源": "306",
    "惠州": "307", "江门": "308", "揭阳": "309", "茂名": "310", "梅州": "311",
    "清远": "312", "汕头": "313", "汕尾": "314", "韶关": "315", "深圳": "316",
    "阳江": "317", "云浮": "318", "湛江": "319", "肇庆": "320", "中山": "321",
    "珠海": "322",
    # 河北
    "石家庄": "653", "保定": "646", "沧州": "647", "承德": "648", "邯郸": "649",
    "衡水": "650", "廊坊": "651", "秦皇岛": "652", "唐山": "654", "邢台": "655",
    "张家口": "656",
    # 山西
    "太原": "1866", "长治": "1859", "大同": "1860", "晋城": "1861", "晋中": "1862",
    "临汾": "1863", "吕梁": "1864", "朔州": "1865", "忻州": "1867", "运城": "1868",
    # 内蒙古
    "呼和浩特": "1566", "阿拉善盟": "1561", "巴彦淖尔市": "1562", "包头": "1563",
    "赤峰": "1564", "鄂尔多斯": "1565", "呼伦贝尔": "1567", "通辽": "1568",
    "乌海": "1569", "乌兰察布": "1570", "锡林郭勒盟": "1571", "兴安盟": "1572",
    # 辽宁
    "沈阳": "1500", "鞍山": "1489", "本溪": "1490", "朝阳": "1491", "大连": "1492",
    "丹东": "1493", "抚顺": "1494", "阜新": "1495", "葫芦岛": "1496", "锦州": "1497",
    "辽阳": "1498", "盘锦": "1499", "铁岭": "1501", "营口": "1502",
    # 吉林
    "长春": "1251", "白城": "1249", "白山": "1250", "吉林": "1252", "辽源": "1253",
    "四平": "1254", "松原": "1255", "通化": "1256", "延边朝鲜族自治州": "1257",
    # 黑龙江
    "哈尔滨": "952", "大庆": "950", "大兴安岭": "951", "鹤岗": "953", "黑河": "954",
    "鸡西": "955", "佳木斯": "956", "牡丹江": "957", "七台河": "958", "齐齐哈尔": "959",
    "双鸭山": "960", "绥化": "961", "伊春": "962",
    # 江苏
    "南京": "1310", "常州": "1307", "淮安": "1308", "连云港": "1309", "南通": "1311",
    "苏州": "1312", "宿迁": "1313", "泰州": "1314", "无锡": "1315", "徐州": "1316",
    "盐城": "1317", "扬州": "1318", "镇江": "1319",
    # 浙江
    "杭州": "2595", "湖州": "2596", "嘉兴": "2597", "金华": "2598", "丽水": "2599",
    "宁波": "2600", "绍兴": "2601", "台州": "2602", "温州": "2603", "舟山": "2604",
    "衢州": "2605",
    # 安徽
    "合肥": "38", "安庆": "32", "蚌埠": "33", "池州": "35", "滁州": "36", "阜阳": "37",
    "淮北": "39", "淮南": "40", "黄山": "41", "六安": "42", "马鞍山": "43", "宿州": "44",
    "铜陵": "45", "芜湖": "46", "宣城": "47", "亳州": "48",
    # 福建
    "福州": "130", "龙岩": "131", "南平": "132", "宁德": "133", "莆田": "134",
    "泉州": "135", "三明": "136", "厦门": "137", "漳州": "138",
    # 江西
    "南昌": "1392", "抚州": "1387", "赣州": "1388", "吉安": "1389", "景德镇": "1390",
    "九江": "1391", "萍乡": "1393", "上饶": "1394", "新余": "1395", "宜春": "1396",
    "鹰潭": "1397",
    # 山东
    "济南": "1738", "滨州": "1734", "德州": "1735", "东营": "1736", "菏泽": "1737",
    "济宁": "1739", "莱芜": "1740", "聊城": "1741", "临沂": "1742", "青岛": "1743",
    "日照": "1744", "泰安": "1745", "威海": "1746", "潍坊": "1747", "烟台": "1748",
    "枣庄": "1749", "淄博": "1750",
    # 河南
    "郑州": "817", "安阳": "804", "鹤壁": "805", "济源": "806", "焦作": "807",
    "开封": "808", "洛阳": "809", "南阳": "810", "平顶山": "811", "三门峡": "812",
    "商丘": "813", "新乡": "814", "信阳": "815", "许昌": "816", "周口": "818",
    "驻马店": "819", "漯河": "820", "濮阳": "821",
    # 湖北
    "武汉": "1052", "鄂州": "1041", "恩施土家族苗族自治州": "1042", "黄冈": "1043",
    "黄石": "1044", "荆门": "1045", "荆州": "1046", "潜江": "1047", "神农架林区": "1048",
    "十堰": "1049", "随州": "1050", "天门": "1051", "仙桃": "1053", "咸宁": "1054",
    "襄阳": "1055", "孝感": "1056", "宜昌": "1057",
    # 湖南
    "长沙": "1135", "常德": "1134", "郴州": "1136", "衡阳": "1137", "怀化": "1138",
    "娄底": "1139", "邵阳": "1140", "湘潭": "1141", "湘西土家族苗族自治州": "1142",
    "益阳": "1143", "永州": "1144", "岳阳": "1145", "张家界": "1146", "株洲": "1147",
    # 广西
    "南宁": "425", "百色": "415", "北海": "416", "崇左": "417", "防城港": "418",
    "桂林": "419", "贵港": "420", "河池": "421", "贺州": "422", "来宾": "423",
    "柳州": "424", "钦州": "426", "梧州": "427", "玉林": "428",
    # 海南
    "海口": "616", "白沙黎族自治县": "610", "保亭黎族苗族自治县": "611",
    "昌江黎族自治县": "612", "澄迈县": "613", "定安县": "614", "东方": "615",
    "乐东黎族自治县": "617", "临高县": "618", "陵水黎族自治县": "619", "琼海": "620",
    "琼中黎族苗族自治县": "621", "三亚": "622", "屯昌县": "623", "万宁": "624",
    "文昌": "625", "五指山": "626", "儋州": "3272", "三沙市": "627",
    # 四川
    "成都": "2087", "阿坝藏族羌族自治州": "2085", "巴中": "2086", "达州": "2088",
    "德阳": "2089", "甘孜藏族自治州": "2090", "广安": "2091", "广元": "2092",
    "乐山": "2093", "凉山彝族自治州": "2094", "眉山": "2095", "绵阳": "2096",
    "南充": "2097", "内江": "2098", "攀枝花": "2099", "遂宁": "2100", "雅安": "2101",
    "宜宾": "2102", "资阳": "2103", "自贡": "2104", "泸州": "2105",
    # 贵州
    "贵阳": "521", "安顺": "519", "毕节": "520", "六盘水": "522",
    "黔东南苗族侗族自治州": "523", "黔南布依族苗族自治州": "524",
    "黔西南布依族苗族自治州": "525", "铜仁": "526", "遵义": "527",
    # 云南
    "昆明": "2460", "保山": "2454", "楚雄彝族自治州": "2455", "大理白族自治州": "2456",
    "德宏傣族景颇族自治州": "2457", "迪庆藏族自治州": "2458",
    "红河哈尼族彝族自治州": "2459", "丽江": "2461", "临沧": "2462",
    "怒江傈傈族自治州": "2463", "曲靖": "2464", "普洱市": "2465",
    "文山壮族苗族自治州": "2466", "西双版纳傣族自治州": "2467", "玉溪": "2468",
    "昭通": "2469",
    # 西藏
    "拉萨": "2270", "阿里": "2268", "昌都": "2269", "林芝": "2271", "那曲": "2272",
    "日喀则": "2273", "山南": "2274",
    # 陕西
    "西安": "1983", "安康": "1977", "宝鸡": "1978", "汉中": "1979", "商洛": "1980",
    "铜川": "1981", "渭南": "1982", "咸阳": "1984", "延安": "1985", "榆林": "1986",
    # 甘肃
    "兰州": "213", "白银": "207", "定西": "208", "甘南藏族自治州": "209", "嘉峪关": "210",
    "金昌": "211", "酒泉": "212", "临夏回族自治州": "214", "陇南": "215", "平凉": "216",
    "庆阳": "217", "天水": "218", "武威": "219", "张掖": "220",
    # 青海
    "西宁": "1692", "果洛藏族自治州": "1686", "海北藏族自治州": "1687", "海东": "1688",
    "海南藏族自治州": "1689", "海西蒙古族藏族自治州": "1690", "黄南藏族自治州": "1691",
    "玉树藏族自治州": "1693",
    # 宁夏
    "银川": "1665", "固原": "1662", "石嘴山": "1663", "吴忠": "1664", "中卫": "1681",
    # 新疆
    "乌鲁木齐": "2361", "阿克苏": "2348", "阿拉尔": "2349", "巴音郭楞蒙古自治州": "2350",
    "博尔塔拉蒙古自治州": "2351", "昌吉回族自治州": "2352", "哈密": "2353", "和田": "2354",
    "喀什": "2355", "克拉玛依": "2356", "克孜勒苏柯尔克孜自治州": "2357", "石河子": "2358",
    "图木舒克": "2359", "吐鲁番": "2360", "五家渠": "2362", "伊犁哈萨克自治州": "2363",
    "塔城地区": "3547", "阿勒泰地区": "3550", "可克达拉市": "3135", "昆玉市": "3136",
    "铁门关市": "3548", "双河市": "3549", "北屯市": "3551",
}

CITY_TO_PROVINCE = {
    # 广东
    "广州": "广东", "潮州": "广东", "东莞": "广东", "佛山": "广东", "河源": "广东",
    "惠州": "广东", "江门": "广东", "揭阳": "广东", "茂名": "广东", "梅州": "广东",
    "清远": "广东", "汕头": "广东", "汕尾": "广东", "韶关": "广东", "深圳": "广东",
    "阳江": "广东", "云浮": "广东", "湛江": "广东", "肇庆": "广东", "中山": "广东",
    "珠海": "广东",
    # 河北
    "石家庄": "河北", "保定": "河北", "沧州": "河北", "承德": "河北", "邯郸": "河北",
    "衡水": "河北", "廊坊": "河北", "秦皇岛": "河北", "唐山": "河北", "邢台": "河北",
    "张家口": "河北",
    # 山西
    "太原": "山西", "长治": "山西", "大同": "山西", "晋城": "山西", "晋中": "山西",
    "临汾": "山西", "吕梁": "山西", "朔州": "山西", "忻州": "山西", "运城": "山西",
    # 内蒙古
    "呼和浩特": "内蒙古", "阿拉善盟": "内蒙古", "巴彦淖尔市": "内蒙古", "包头": "内蒙古",
    "赤峰": "内蒙古", "鄂尔多斯": "内蒙古", "呼伦贝尔": "内蒙古", "通辽": "内蒙古",
    "乌海": "内蒙古", "乌兰察布": "内蒙古", "锡林郭勒盟": "内蒙古", "兴安盟": "内蒙古",
    # 辽宁
    "沈阳": "辽宁", "鞍山": "辽宁", "本溪": "辽宁", "朝阳": "辽宁", "大连": "辽宁",
    "丹东": "辽宁", "抚顺": "辽宁", "阜新": "辽宁", "葫芦岛": "辽宁", "锦州": "辽宁",
    "辽阳": "辽宁", "盘锦": "辽宁", "铁岭": "辽宁", "营口": "辽宁",
    # 吉林
    "长春": "吉林", "白城": "吉林", "白山": "吉林", "吉林": "吉林", "辽源": "吉林",
    "四平": "吉林", "松原": "吉林", "通化": "吉林", "延边朝鲜族自治州": "吉林",
    # 黑龙江
    "哈尔滨": "黑龙江", "大庆": "黑龙江", "大兴安岭": "黑龙江", "鹤岗": "黑龙江",
    "黑河": "黑龙江", "鸡西": "黑龙江", "佳木斯": "黑龙江", "牡丹江": "黑龙江",
    "七台河": "黑龙江", "齐齐哈尔": "黑龙江", "双鸭山": "黑龙江", "绥化": "黑龙江",
    "伊春": "黑龙江",
    # 江苏
    "南京": "江苏", "常州": "江苏", "淮安": "江苏", "连云港": "江苏", "南通": "江苏",
    "苏州": "江苏", "宿迁": "江苏", "泰州": "江苏", "无锡": "江苏", "徐州": "江苏",
    "盐城": "江苏", "扬州": "江苏", "镇江": "江苏",
    # 浙江
    "杭州": "浙江", "湖州": "浙江", "嘉兴": "浙江", "金华": "浙江", "丽水": "浙江",
    "宁波": "浙江", "绍兴": "浙江", "台州": "浙江", "温州": "浙江", "舟山": "浙江",
    "衢州": "浙江",
    # 安徽
    "合肥": "安徽", "安庆": "安徽", "蚌埠": "安徽", "池州": "安徽", "滁州": "安徽",
    "阜阳": "安徽", "淮北": "安徽", "淮南": "安徽", "黄山": "安徽", "六安": "安徽",
    "马鞍山": "安徽", "宿州": "安徽", "铜陵": "安徽", "芜湖": "安徽", "宣城": "安徽",
    "亳州": "安徽",
    # 福建
    "福州": "福建", "龙岩": "福建", "南平": "福建", "宁德": "福建", "莆田": "福建",
    "泉州": "福建", "三明": "福建", "厦门": "福建", "漳州": "福建",
    # 江西
    "南昌": "江西", "抚州": "江西", "赣州": "江西", "吉安": "江西", "景德镇": "江西",
    "九江": "江西", "萍乡": "江西", "上饶": "江西", "新余": "江西", "宜春": "江西",
    "鹰潭": "江西",
    # 山东
    "济南": "山东", "滨州": "山东", "德州": "山东", "东营": "山东", "菏泽": "山东",
    "济宁": "山东", "莱芜": "山东", "聊城": "山东", "临沂": "山东", "青岛": "山东",
    "日照": "山东", "泰安": "山东", "威海": "山东", "潍坊": "山东", "烟台": "山东",
    "枣庄": "山东", "淄博": "山东",
    # 河南
    "郑州": "河南", "安阳": "河南", "鹤壁": "河南", "济源": "河南", "焦作": "河南",
    "开封": "河南", "洛阳": "河南", "南阳": "河南", "平顶山": "河南", "三门峡": "河南",
    "商丘": "河南", "新乡": "河南", "信阳": "河南", "许昌": "河南", "周口": "河南",
    "驻马店": "河南", "漯河": "河南", "濮阳": "河南",
    # 湖北
    "武汉": "湖北", "鄂州": "湖北", "恩施土家族苗族自治州": "湖北", "黄冈": "湖北",
    "黄石": "湖北", "荆门": "湖北", "荆州": "湖北", "潜江": "湖北", "神农架林区": "湖北",
    "十堰": "湖北", "随州": "湖北", "天门": "湖北", "仙桃": "湖北", "咸宁": "湖北",
    "襄阳": "湖北", "孝感": "湖北", "宜昌": "湖北",
    # 湖南
    "长沙": "湖南", "常德": "湖南", "郴州": "湖南", "衡阳": "湖南", "怀化": "湖南",
    "娄底": "湖南", "邵阳": "湖南", "湘潭": "湖南", "湘西土家族苗族自治州": "湖南",
    "益阳": "湖南", "永州": "湖南", "岳阳": "湖南", "张家界": "湖南", "株洲": "湖南",
    # 广西
    "南宁": "广西", "百色": "广西", "北海": "广西", "崇左": "广西", "防城港": "广西",
    "桂林": "广西", "贵港": "广西", "河池": "广西", "贺州": "广西", "来宾": "广西",
    "柳州": "广西", "钦州": "广西", "梧州": "广西", "玉林": "广西",
    # 海南
    "海口": "海南", "白沙黎族自治县": "610", "保亭黎族苗族自治县": "611",
    "昌江黎族自治县": "612", "澄迈县": "613", "定安县": "614", "东方": "615",
    "乐东黎族自治县": "617", "临高县": "618", "陵水黎族自治县": "619", "琼海": "620",
    "琼中黎族苗族自治县": "621", "三亚": "622", "屯昌县": "623", "万宁": "624",
    "文昌": "625", "五指山": "626", "儋州": "3272", "三沙市": "627",
    # 四川
    "成都": "四川", "阿坝藏族羌族自治州": "四川", "巴中": "四川", "达州": "四川",
    "德阳": "四川", "甘孜藏族自治州": "四川", "广安": "四川", "广元": "四川",
    "乐山": "四川", "凉山彝族自治州": "四川", "眉山": "四川", "绵阳": "四川",
    "南充": "四川", "内江": "四川", "攀枝花": "四川", "遂宁": "四川", "雅安": "四川",
    "宜宾": "四川", "资阳": "四川", "自贡": "四川", "泸州": "四川",
    # 贵州
    "贵阳": "贵州", "安顺": "贵州", "毕节": "贵州", "六盘水": "贵州",
    "黔东南苗族侗族自治州": "贵州", "黔南布依族苗族自治州": "贵州",
    "黔西南布依族苗族自治州": "贵州", "铜仁": "贵州", "遵义": "贵州",
    # 云南
    "昆明": "云南", "保山": "云南", "楚雄彝族自治州": "云南", "大理白族自治州": "云南",
    "德宏傣族景颇族自治州": "云南", "迪庆藏族自治州": "云南",
    "红河哈尼族彝族自治州": "云南", "丽江": "云南", "临沧": "云南",
    "怒江傈傈族自治州": "云南", "曲靖": "云南", "普洱市": "云南",
    "文山壮族苗族自治州": "云南", "西双版纳傣族自治州": "云南", "玉溪": "云南",
    "昭通": "云南",
    # 西藏
    "拉萨": "西藏", "阿里": "西藏", "昌都": "西藏", "林芝": "西藏", "那曲": "西藏",
    "日喀则": "西藏", "山南": "西藏",
    # 陕西
    "西安": "陕西", "安康": "陕西", "宝鸡": "陕西", "汉中": "陕西", "商洛": "陕西",
    "铜川": "陕西", "渭南": "陕西", "咸阳": "陕西", "延安": "陕西", "榆林": "陕西",
    # 甘肃
    "兰州": "甘肃", "白银": "甘肃", "定西": "甘肃", "甘南藏族自治州": "甘肃",
    "嘉峪关": "甘肃", "金昌": "甘肃", "酒泉": "甘肃", "临夏回族自治州": "甘肃",
    "陇南": "甘肃", "平凉": "甘肃", "庆阳": "甘肃", "天水": "甘肃", "武威": "甘肃",
    "张掖": "甘肃",
    # 青海
    "西宁": "青海", "果洛藏族自治州": "青海", "海北藏族自治州": "青海", "海东": "青海",
    "海南藏族自治州": "青海", "海西蒙古族藏族自治州": "青海", "黄南藏族自治州": "青海",
    "玉树藏族自治州": "青海",
    # 宁夏
    "银川": "宁夏", "固原": "宁夏", "石嘴山": "宁夏", "吴忠": "宁夏", "中卫": "宁夏",
    # 新疆
    "乌鲁木齐": "新疆", "阿克苏": "新疆", "阿拉尔": "新疆", "巴音郭楞蒙古自治州": "新疆",
    "博尔塔拉蒙古自治州": "新疆", "昌吉回族自治州": "新疆", "哈密": "新疆", "和田": "新疆",
    "喀什": "新疆", "克拉玛依": "新疆", "克孜勒苏柯尔克孜自治州": "新疆", "石河子": "新疆",
    "图木舒克": "新疆", "吐鲁番": "新疆", "五家渠": "新疆", "伊犁哈萨克自治州": "新疆",
    "塔城地区": "新疆", "阿勒泰地区": "新疆", "可克达拉市": "新疆", "昆玉市": "新疆",
    "铁门关市": "新疆", "双河市": "新疆", "北屯市": "新疆",
}

# 构建 ID -> 省份 映射
ID_TO_PROVINCE = {}
for province, pid in PROVINCE_ID_MAP.items():
    ID_TO_PROVINCE[pid] = province
for city, cid in CITY_ID_MAP.items():
    ID_TO_PROVINCE[cid] = CITY_TO_PROVINCE.get(city, "未知")
ID_TO_PROVINCE["2703"] = "全国"

# 预定义的南区省份组合
SOUTH_PROVINCES = ['广东', '广西', '海南', '湖南', '江西']
SOUTH_IDS = [PROVINCE_ID_MAP[p] for p in SOUTH_PROVINCES]

# ==================== 其他配置 ====================
TIME_OPTION_MAP = {"2": "1", "3": "2", "4": "5", "5": "3"}

# ==================== 关键词定义（从配置文件加载）====================
import json
import os

DEFAULT_KEYWORDS = {
    "工程类": [
        "工程", "建设", "施工", "改造", "装修", "修缮", "环境", "安装", "机房", "实验室",
        "改造工程", "建设工程", "建筑安装", "装饰装修", "净化工程", "屏蔽工程", "防护工程",
        "改建", "扩建", "修建", "土木工程", "电力工程", "给排水", "暖通", "消防工程", "搬迁"
    ],
    "服务类": [
        "维保", "维修", "保养", "托管", "租赁", "服务", "保修", "维护", "售后", "线圈",
        "技术支持", "运营", "外包", "派遣", "劳务", "软件", "记录本", "培训", "维护保养",
        "全保", "球管", "核磁管", "探测器", "高压发生器", "回收", "PACS", "报废", "转让"
    ],
    "仪器类": [
        "超声", "仪器", "麻醉机", "监护", "内镜", "内窥镜", "波谱仪", "新生儿科", "检验科",
        "激光治疗仪", "手术床", "急救和生命支持设备", "血液透析", "口腔", "耳鼻喉", "鼻咽", "工作站",
        "分析仪", "检测仪", "诊断仪", "显微镜", "离心机", "灭菌器", "消毒设备", "呼吸机",
        "除颤仪", "心电图机", "监护仪", "注射泵", "输液泵", "病床", "无影灯", "吊塔", "高压注射器"
    ],
    "CT类": ["CT", "计算机断层扫描", "计算机体层摄影"],
    "DSA类": ["DSA", "数字减影", "血管机", "血管造影"],
    "MR类": ["MR", "磁共振", "核磁"],
    "其他": ["医用X线", "医用设备", "医疗设备", "设备更新", "医共体", "卫生健康"]
}

def load_keywords_config(config_file='keywords_config.json'):
    """加载关键词配置文件，如果不存在则创建默认配置并返回"""
    if not os.path.exists(config_file):
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_KEYWORDS, f, ensure_ascii=False, indent=2)
        print(f"未找到关键词配置文件，已自动创建默认文件：{config_file}")
        return DEFAULT_KEYWORDS
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        # 确保所有必要分类都存在（防止用户删减）
        for key in DEFAULT_KEYWORDS:
            if key not in config:
                config[key] = DEFAULT_KEYWORDS[key]
        return config
    except Exception as e:
        print(f"读取配置文件失败：{e}，使用默认关键词。")
        return DEFAULT_KEYWORDS

# 加载关键词配置
keywords_config = load_keywords_config()
KEYWORDS_CONSTRUCTION = keywords_config["工程类"]
KEYWORDS_SERVICE = keywords_config["服务类"]
KEYWORDS_INSTRUMENT = keywords_config["仪器类"]
KEYWORDS_CT = keywords_config["CT类"]
KEYWORDS_DSA = keywords_config["DSA类"]
KEYWORDS_MR = keywords_config["MR类"]
KEYWORDS_OTHER = keywords_config["其他"]

BRAND_LIST = ['飞利浦', '西门子', 'GE', '联影', '东软', '佳能', '东芝', '日立', '万东']

CONTENT_TYPE_DISPLAY_MAP = {'1': '招标信息', '2': '中标信息', '3': '采购意向'}
CONTENT_TYPE_API_MAP = {'招标信息': 'zhaobiao', '中标信息': 'zhongbiao', '采购意向': 'caigou'}


# ==================== 工具函数 ====================
def set_console_window_top():
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.SetWindowPos(hwnd, -1, 0, 0, 0, 0, 0x0002 | 0x0001)
            ctypes.windll.user32.InvalidateRect(hwnd, None, True)
            ctypes.windll.user32.UpdateWindow(hwnd)
            ctypes.windll.user32.ShowWindow(hwnd, 9)
            ctypes.windll.user32.SetForegroundWindow(hwnd)
        winsound.MessageBeep()
    except:
        pass

def parse_date_input(date_str):
    if not date_str:
        return None
    patterns = [
        r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})',
        r'(\d{4})(\d{2})(\d{2})'
    ]
    for pat in patterns:
        m = re.search(pat, date_str)
        if m:
            year, month, day = m.groups()
            month = month.zfill(2)
            day = day.zfill(2)
            try:
                datetime(int(year), int(month), int(day))
                return f"{year}/{month}/{day}"
            except ValueError:
                continue
    return None

def parse_year_month(ym_str):
    """解析用户输入的年月，返回 YYYY/MM 格式"""
    if not ym_str:
        return None
    patterns = [
        r'(\d{4})[-/年](\d{1,2})',  # 2026-01, 2026/01, 2026年1
        r'(\d{4})(\d{2})',           # 202601
    ]
    for pat in patterns:
        m = re.search(pat, ym_str)
        if m:
            year, month = m.groups()
            month = month.zfill(2)
            if 1 <= int(month) <= 12:
                return f"{year}/{month}"
    return None

def get_current_year_month():
    """返回当前年份和月份元组 (year, month)"""
    now = datetime.now()
    return now.year, now.month

def get_three_months_later():
    """返回当前月份后3个月的年份和月份元组 (year, month)"""
    now = datetime.now()
    year = now.year
    month = now.month + 3
    if month > 12:
        year += (month - 1) // 12
        month = (month - 1) % 12 + 1
    return year, month

def get_date_range(option, custom_days=None):
    end_date = datetime.now()
    if option == '1':
        start_date = end_date
    elif option == '6' and custom_days is not None:
        start_date = end_date - timedelta(days=int(custom_days))
    else:
        return None, None
    return start_date.strftime('%Y/%m/%d'), end_date.strftime('%Y/%m/%d')

def convert_fuzzy_time(time_str):
    if not isinstance(time_str, str):
        return time_str
    if any(key in time_str for key in ['小时前', '分钟前', '天前']):
        today = datetime.now().strftime('%Y-%m-%d')
        return f"{today} {time_str}"
    return time_str

# ========== 内容类型判断函数 ==========
# ========== 重构后的内容类型判断函数 ==========
def determine_content_type(title, detail_text=''):
    """
    判断内容类型，按优先级：
    1. 标题中包含工程类、服务类、仪器类关键词 → 返回对应类别（单一类别，按工程→服务→仪器顺序匹配）。
    2. 标题中包含CT类、DSA类、MR类关键词 → 返回所有匹配的类别（用逗号+空格分隔）。
    3. 若标题未命中以上，则从公告详情中搜索CT/DSA/MR关键词，若命中则返回所有匹配的类别。
    4. 若详情中仍未命中设备类，则从详情中搜索工程/服务/仪器关键词，返回第一个匹配的类别。
    5. 若均未命中，返回'其他'。
    """
    title_lower = title.lower()

    # 第一优先级：标题中的工程、服务、仪器（只返回第一个匹配到的类别）
    if any(kw.lower() in title_lower for kw in KEYWORDS_CONSTRUCTION):
        return '工程类'
    if any(kw.lower() in title_lower for kw in KEYWORDS_SERVICE):
        return '服务类'
    if any(kw.lower() in title_lower for kw in KEYWORDS_INSTRUMENT):
        return '仪器类'

    # 第二优先级：标题中的设备类（可能多个）
    matched_title = []
    if any(kw.lower() in title_lower for kw in KEYWORDS_CT):
        matched_title.append('CT类')
    if any(kw.lower() in title_lower for kw in KEYWORDS_DSA):
        matched_title.append('DSA类')
    if any(kw.lower() in title_lower for kw in KEYWORDS_MR):
        matched_title.append('MR类')
    if matched_title:
        return ', '.join(matched_title)

    # 第三优先级：详情中的设备类（可能多个）
    if detail_text:
        detail_lower = detail_text.lower()
        matched_detail = []
        if any(kw.lower() in detail_lower for kw in KEYWORDS_CT):
            matched_detail.append('CT类')
        if any(kw.lower() in detail_lower for kw in KEYWORDS_DSA):
            matched_detail.append('DSA类')
        if any(kw.lower() in detail_lower for kw in KEYWORDS_MR):
            matched_detail.append('MR类')
        if matched_detail:
            return ', '.join(matched_detail)

        # 第四优先级：详情中的工程、服务、仪器（只返回第一个匹配到的类别）
        if any(kw.lower() in detail_lower for kw in KEYWORDS_CONSTRUCTION):
            return '工程类'
        if any(kw.lower() in detail_lower for kw in KEYWORDS_SERVICE):
            return '服务类'
        if any(kw.lower() in detail_lower for kw in KEYWORDS_INSTRUMENT):
            return '仪器类'

    # 第五优先级：其他
    return '其他'

# ==================== Cookie 获取（支持重启时关闭浏览器）====================
def get_cookies_from_manual_login():
    global global_driver

    # 如果之前有浏览器实例，先关闭
    if global_driver:
        try:
            global_driver.quit()
        except:
            pass
        global_driver = None

    # 配置浏览器选项以抑制日志
    edge_options = EdgeOptions()
    edge_options.add_argument("--log-level=3")
    edge_options.add_argument("--silent")
    edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])

    chrome_options = ChromeOptions()
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--silent")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

    firefox_options = FirefoxOptions()
    firefox_options.add_argument("--log-level=3")  # Firefox 可能不支持，但无害

    safari_options = SafariOptions()  # Safari 无需特殊配置

    browsers_to_try = [
        ('Chrome', webdriver.Chrome, chrome_options),
        ('Edge', webdriver.Edge, edge_options),
        ('Firefox', webdriver.Firefox, firefox_options),
        ('Safari', webdriver.Safari, safari_options)
    ]

    driver = None
    browser_name = None
    for name, driver_class, options in browsers_to_try:
        try:
            print(f"正在尝试启动 {name} 浏览器...")
            driver = driver_class(options=options)
            browser_name = name
            break
        except Exception as e:
            print(f"启动 {name} 失败：{e}")
            continue

    if driver is None:
        print("错误：无法启动任何支持的浏览器。请安装 Chrome、Edge、Firefox，或在 macOS 上启用 Safari 远程自动化（执行 'safaridriver --enable'）。")
        sys.exit(1)

    print(f"成功启动 {browser_name} 浏览器（由 Selenium Manager 管理驱动）。")
    driver.get("https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/search")
    print("\n" + "="*50)
    print("请在浏览器中手动登录乙方宝网站。")
    print("登录成功后，请回到此窗口按回车键继续...")
    print("注意：浏览器将保持打开状态，您可以直接在此浏览器中新建标签页手工验证公告。")
    input_with_restart("按回车键继续")

    selenium_cookies = driver.get_cookies()
    global_driver = driver

    cj = requests.cookies.RequestsCookieJar()
    for cookie in selenium_cookies:
        cj.set(cookie['name'], cookie['value'], domain=cookie.get('domain', ''))

    with open(COOKIE_FILE, 'wb') as f:
        pickle.dump(cj, f)
    print(f"Cookie 已保存到 {COOKIE_FILE}，后续运行可直接使用。")
    return cj

def load_cookies():
    if os.path.exists(COOKIE_FILE):
        try:
            with open(COOKIE_FILE, 'rb') as f:
                cj = pickle.load(f)
            test_url = BASE_API_URL + "?pageSize=1&pageNum=1&pageFrom=zhaobiao&keyword=test"
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }
            r = requests.get(test_url, cookies=cj, headers=headers, timeout=5)
            if r.status_code == 200:
                data = r.json()
                if data.get('code') == 200:
                    print("从本地文件加载 Cookie 成功。")
                    return cj
                else:
                    print("本地 Cookie 已失效，需要重新登录。")
            else:
                print("本地 Cookie 可能已失效，需要重新登录。")
        except Exception as e:
            print(f"加载本地 Cookie 失败：{e}，将启动浏览器重新获取。")
    else:
        print("未找到保存的 Cookie 文件，需要手动登录。")

    return get_cookies_from_manual_login()


# ==================== API 构建 ====================
def build_api_url(province_ids, keyword, content_type, search_type, time_option,
                  page_num=1, start_date=None, end_date=None,
                  purchase_time_type=None, purchase_times=None):
    params = {
        "pageSize": 30,
        "pageNum": page_num,
        "pageFrom": content_type,
        "keyword": keyword,
        "queryType": "",
        "offSet": "",
        "areaIds": province_ids,
        "times": "",
        "searchType": "2" if search_type == "标题检索" else "1",
        "nature": "",
        "defTimeFlag": "0",
    }
    if content_type == 'caigou':
        params["filterCondition"] = "1"
    else:
        params["filterCondition"] = "2"

    if start_date and end_date:
        params["times"] = f"{start_date},{end_date}"
        params["defTimeFlag"] = "1"
        params["timeOption"] = ""
    else:
        params["timeOption"] = time_option or ""

    if content_type == 'caigou':
        if purchase_times:
            params["times"] = purchase_times
            params["defTimeFlag"] = "1"
            params["purchaseTimeType"] = ""
        elif purchase_time_type is not None:
            params["purchaseTimeType"] = str(purchase_time_type)
            if "times" in params and not params["times"]:
                del params["times"]
            params["defTimeFlag"] = "0"
        else:
            params["purchaseTimeType"] = "-1"

    params = {k: v for k, v in params.items() if v is not None}
    return f"{BASE_API_URL}?{urlencode(params)}"


def fetch_search_results(url, retries=5):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*", "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive", "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://qiye.qianlima.com", "Referer": "https://qiye.qianlima.com/new_qd_yfbsite/",
        "Sec-Fetch-Dest": "empty", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Site": "same-origin",
    }
    for attempt in range(retries):
        try:
            time.sleep(random.uniform(1, 3))
            response = requests.get(url, cookies=cj, headers=headers, timeout=10)
            print(f"  响应状态码: {response.status_code}")
            if response.status_code == 200:
                data = response.json()
                if data.get("code") == 200:
                    return data.get("data", {})
                else:
                    print(f"  API返回错误: {data.get('msg')}，完整响应：{data}")
                    if "认证失败" in data.get('msg', ''):
                        print("  Cookie 可能已过期，请重新运行程序以手动登录。")
                        sys.exit(1)
                    elif data.get('code') == 500:
                        wait_time = (attempt + 1) * 10
                        print(f"  服务器返回500，等待 {wait_time} 秒后重试...")
                        time.sleep(wait_time)
                        continue
            else:
                print(f"  HTTP错误: {response.status_code}，响应内容：{response.text[:200]}")
        except Exception as e:
            print(f"  请求异常: {e}")
        time.sleep(2)
    print("  已达到最大重试次数，放弃该请求")
    return None


# ==================== 纯公告页面解析 ====================
def parse_share_page_for_zhaobiao(share_url):
    """从招标公告的纯公告页面解析标的名称和数量。返回 (item_names, quantities)"""
    if not share_url:
        return '', ''
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = requests.get(share_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return '', ''
        soup = BeautifulSoup(resp.text, 'html.parser')
        item_keywords = ['采购标的', '标的名称', '货物名称']
        qty_keywords = ['数量']
        item_names = []
        quantities = []
        for table in soup.find_all('table'):
            header_row = table.find('tr')
            if not header_row:
                continue
            header_cells = header_row.find_all(['th', 'td'])
            header_texts = [cell.get_text(strip=True) for cell in header_cells]
            if not any(any(kw in h for kw in item_keywords) for h in header_texts):
                continue
            if not any(any(kw in h for kw in qty_keywords) for h in header_texts):
                continue
            item_col = None
            qty_col = None
            for i, text in enumerate(header_texts):
                if any(kw in text for kw in item_keywords):
                    item_col = i
                if any(kw in text for kw in qty_keywords):
                    qty_col = i
            if item_col is None or qty_col is None:
                continue
            rows = table.find_all('tr')[1:]
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) <= max(item_col, qty_col):
                    continue
                item = cells[item_col].get_text(strip=True)
                qty = cells[qty_col].get_text(strip=True)
                if item:
                    item_names.append(item)
                if qty:
                    qty_match = re.search(r'(\d+(\.\d+)?)', qty)
                    quantities.append(qty_match.group(1) if qty_match else qty)
        return '；'.join(item_names), '；'.join(quantities)
    except Exception as e:
        print(f"  招标纯公告解析出错: {e}")
        return '', ''


def parse_share_page_for_zhongbiao(share_url):
    """从中标公告的纯公告页面解析标的名称、品牌、规格型号、数量、单价。"""
    if not share_url:
        return '', '', '', '', ''
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = requests.get(share_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return '', '', '', '', ''
        soup = BeautifulSoup(resp.text, 'html.parser')
        target_keywords = ['货物名称', '品牌', '规格型号', '数量', '单价']
        item_names, brands, specs, quantities, prices = [], [], [], [], []
        for table in soup.find_all('table'):
            header_row = table.find('tr')
            if not header_row:
                continue
            header_cells = header_row.find_all(['th', 'td'])
            header_texts = [cell.get_text(strip=True) for cell in header_cells]
            matched = sum(1 for kw in target_keywords if any(kw in h for h in header_texts))
            if matched < 3:
                continue
            col_idx = {}
            for i, text in enumerate(header_texts):
                if '货物名称' in text or '标的名称' in text or '采购标的' in text:
                    col_idx['item'] = i
                elif '品牌' in text:
                    col_idx['brand'] = i
                elif '规格型号' in text or '规格' in text:
                    col_idx['spec'] = i
                elif '数量' in text:
                    col_idx['qty'] = i
                elif '单价' in text:
                    col_idx['price'] = i
            rows = table.find_all('tr')[1:]
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if not cells:
                    continue
                if 'item' in col_idx and col_idx['item'] < len(cells):
                    val = cells[col_idx['item']].get_text(strip=True)
                    if val:
                        item_names.append(val)
                if 'brand' in col_idx and col_idx['brand'] < len(cells):
                    val = cells[col_idx['brand']].get_text(strip=True)
                    if val:
                        brands.append(val)
                if 'spec' in col_idx and col_idx['spec'] < len(cells):
                    val = cells[col_idx['spec']].get_text(strip=True)
                    if val:
                        specs.append(val)
                if 'qty' in col_idx and col_idx['qty'] < len(cells):
                    qty_text = cells[col_idx['qty']].get_text(strip=True)
                    qty_match = re.search(r'(\d+(\.\d+)?)', qty_text)
                    quantities.append(qty_match.group(1) if qty_match else qty_text)
                if 'price' in col_idx and col_idx['price'] < len(cells):
                    price_text = cells[col_idx['price']].get_text(strip=True)
                    price_match = re.search(r'[\d,]+(?:\.\d+)?', price_text)
                    prices.append(price_match.group().replace(',', '') if price_match else price_text)
        return ('；'.join(item_names), '；'.join(brands), '；'.join(specs),
                '；'.join(quantities), '；'.join(prices))
    except Exception as e:
        print(f"  中标纯公告解析出错: {e}")
        return '', '', '', '', ''


# ==================== 详情页解析 ====================
def extract_plain_text(html_content):
    if not html_content:
        return ''
    soup = BeautifulSoup(html_content, 'html.parser')
    return soup.get_text(separator='\n', strip=True)


def extract_attachments(detail_data):
    downlink_list = detail_data.get('downlinkList', [])
    if not downlink_list:
        return ''
    attachments = []
    for item in downlink_list:
        title = item.get('title', '').strip()
        url = item.get('url', '').strip()
        if title and url:
            attachments.append(f"{title}：{url}")
    return '\n'.join(attachments)


def fetch_detail_from_share_url(share_url):
    """从分享页面提取采购意向的补充字段"""
    result = {'采购品目': '', '采购需求概况': '', '预算金额': '', '预计采购日期': ''}
    if not share_url:
        return result
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        response = requests.get(share_url, headers=headers, timeout=10)
        if response.status_code != 200:
            return result
        soup = BeautifulSoup(response.text, 'html.parser')
        text = soup.get_text()
        cat_match = re.search(r'采购品目[：:]\s*([^\n\r]+)', text)
        if cat_match:
            result['采购品目'] = cat_match.group(1).strip()
        need_patterns = [r'采购需求概况[：:]\s*([^\n\r]+)', r'主要功能或目标[：:]\s*([^\n\r]+)',
                         r'需满足的要求[：:]\s*([^\n\r]+)', r'采购需求[：:]\s*([^\n\r]+)']
        for pat in need_patterns:
            need_match = re.search(pat, text)
            if need_match:
                result['采购需求概况'] = need_match.group(1).strip()
                break
        budget_match = re.search(r'预算金额[：:]\s*([\d,]+(?:\.\d+)?[万元元]*)', text)
        if budget_match:
            result['预算金额'] = budget_match.group(1).strip()
        else:
            budget_match2 = re.search(r'预算[：:]\s*([\d,]+(?:\.\d+)?[万元元]*)', text)
            if budget_match2:
                result['预算金额'] = budget_match2.group(1).strip()
        date_patterns = [r'预计采购时间[：:]\s*([^\n\r]+)', r'预计采购日期[：:]\s*([^\n\r]+)',
                         r'采购时间[：:]\s*([^\n\r]+)']
        for pat in date_patterns:
            date_match = re.search(pat, text)
            if date_match:
                result['预计采购日期'] = date_match.group(1).strip()
                break
        return result
    except Exception as e:
        print(f"  从分享页面提取信息时出错: {e}")
        return result


def fetch_zhaobiao_detail_api(content_id, area_id):
    url = f"{DETAIL_API_COMMON}?contentId={content_id}&areaId={area_id}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://qiye.qianlima.com",
        "Referer": f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{content_id}/{area_id}",
    }
    try:
        response = requests.get(url, cookies=cj, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data.get("code") == 200:
                d = data.get("data", {})
                summary = d.get('summary', {})
                title = re.sub(r'<[^>]+>', '', d.get('title', ''))
                announce_type = d.get('type', '')
                update_date = d.get('updateDate', '')
                area_name = d.get('areaName', '')
                if area_name:
                    parts = area_name.split('-')
                    province = parts[0] if len(parts) > 0 else ''
                    city = parts[1] if len(parts) > 1 else ''
                else:
                    province, city = '', ''
                share_url = d.get('shareUrl', '')
                hospital = d.get('invitedBidCompany', '')
                bid_end_date = summary.get('bidEndDate', '')
                item_no = summary.get('itemNo', '')
                estimated_amount = summary.get('estimatedAmount', '')
                win_bid_amount = summary.get('winBidAmount', '')
                html = d.get('content', '')
                plain_text = extract_plain_text(html)
                attachments = extract_attachments(d)
                item_names, quantities = parse_share_page_for_zhaobiao(share_url)
                return {
                    '标题': title, '公告类型': announce_type, '发布时间': update_date,
                    '省份': province, '城市': city, '纯公告链接': share_url,
                    '医院名称': hospital, '开标时间': bid_end_date, '采购编号': item_no,
                    '标的名称': item_names, '数量': quantities,
                    '预算金额': estimated_amount, '最高限价': win_bid_amount,
                    '公告详情': plain_text, '附件': attachments,
                }
            else:
                print(f"  招标详情API返回错误: {data.get('msg')}")
        else:
            print(f"  招标详情API HTTP错误: {response.status_code}")
    except Exception as e:
        print(f"  招标详情API请求异常: {e}")
    return {}


def fetch_zhongbiao_detail_api(content_id, area_id):
    url = f"{DETAIL_API_COMMON}?contentId={content_id}&areaId={area_id}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://qiye.qianlima.com",
        "Referer": f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{content_id}/{area_id}",
    }
    try:
        response = requests.get(url, cookies=cj, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data.get("code") == 200:
                d = data.get("data", {})
                summary = d.get('summary', {})
                title = re.sub(r'<[^>]+>', '', d.get('title', ''))
                announce_type = d.get('type', '')
                update_date = d.get('updateDate', '')
                area_name = d.get('areaName', '')
                if area_name:
                    parts = area_name.split('-')
                    province = parts[0] if len(parts) > 0 else ''
                    city = parts[1] if len(parts) > 1 else ''
                else:
                    province, city = '', ''
                share_url = d.get('shareUrl', '')
                hospital = d.get('invitedBidCompany', '')
                item_no = summary.get('itemNo', '')
                estimated_amount = summary.get('estimatedAmount', '')
                win_bid_amount = summary.get('winBidAmount', '')
                supplier = summary.get('winBidUnit', {}).get('unitName', '')
                html = d.get('content', '')
                plain_text = extract_plain_text(html)
                attachments = extract_attachments(d)
                item_names, brands, specs, quantities, prices = parse_share_page_for_zhongbiao(share_url)
                return {
                    '标题': title, '公告类型': announce_type, '发布时间': update_date,
                    '省份': province, '城市': city, '纯公告链接': share_url,
                    '医院名称': hospital, '采购编号': item_no,
                    '标的名称': item_names, '品牌': brands, '规格型号': specs,
                    '数量': quantities, '单价': prices,
                    '预算金额': estimated_amount, '中标金额': win_bid_amount,
                    '供应商名称': supplier, '公告详情': plain_text, '附件': attachments,
                }
            else:
                print(f"  中标详情API返回错误: {data.get('msg')}")
        else:
            print(f"  中标详情API HTTP错误: {response.status_code}")
    except Exception as e:
        print(f"  中标详情API请求异常: {e}")
    return {}


def fetch_caigou_detail_api(content_id, area_id):
    url = f"{DETAIL_API_COMMON}?contentId={content_id}&areaId={area_id}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*", "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://qiye.qianlima.com",
        "Referer": f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{content_id}/{area_id}",
    }
    try:
        response = requests.get(url, cookies=cj, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data.get("code") == 200:
                d = data.get("data", {})
                summary = d.get('summary', {})
                hospital = d.get('invitedBidCompany', '') or summary.get('biddingUnit', {}).get('unitName', '')
                item_name = summary.get('itemNo', '') or d.get('title', '')
                budget = summary.get('estimatedAmount', '')
                plain_text = extract_plain_text(d.get('content', ''))
                attachments = extract_attachments(d)
                share_url = d.get('shareUrl', '')
                share_info = fetch_detail_from_share_url(share_url) if share_url else {}
                return {
                    '医院名称': hospital, '开标时间': '', '标的名称': item_name,
                    '数量': '', '最高限价': '', '预算金额': budget, '预采时间': '',
                    '品牌': '', '规格型号': '', '单价': '', '中标金额': '', '供应商名称': '',
                    '公告详情': plain_text, '附件': attachments,
                    '采购项目名称': '', '采购品目': share_info.get('采购品目', ''),
                    '采购需求概况': share_info.get('采购需求概况', ''),
                    '预计采购日期': share_info.get('预计采购日期', ''),
                }
            else:
                print(f"  采购意向详情API返回错误: {data.get('msg')}")
        else:
            print(f"  采购意向详情API HTTP错误: {response.status_code}")
    except Exception as e:
        print(f"  采购意向详情API请求异常: {e}")
    return {}


def extract_detail_info(detail_url, content_type):
    m = re.search(r'infoDetail/(\d+)/(\d+)', detail_url)
    if not m:
        print(f"  无法从URL提取ID: {detail_url}")
        return {}
    cid, aid = m.groups()
    if content_type == 'zhongbiao':
        return fetch_zhongbiao_detail_api(cid, aid)
    elif content_type == 'caigou':
        return fetch_caigou_detail_api(cid, aid)
    else:
        return fetch_zhaobiao_detail_api(cid, aid)


def is_fuzzy_time(time_str):
    if not isinstance(time_str, str):
        return False
    return '小时前' in time_str or '分钟前' in time_str or '天前' in time_str


# ==================== Excel 保存 ====================
def save_multi_sheet_to_excel(data_dict, file_name, mode='new'):
    full = os.path.join(OUTPUT_DIR, file_name)
    if mode == 'new' or not os.path.exists(full):
        with pd.ExcelWriter(full, engine='openpyxl') as writer:
            for name, df in data_dict.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=name, index=False)
        print(f"数据已保存至新文件: {full}")
    else:
        wb = None
        while wb is None:
            try:
                wb = load_workbook(full)
            except PermissionError:
                print(f"文件 {file_name} 被占用，请关闭后按回车重试...")
                input_with_restart("按回车重试...")
            except Exception as e:
                print(f"打开文件失败：{e}")
                return

        for name, df in data_dict.items():
            if df.empty:
                continue
            if name in wb.sheetnames:
                ws = wb[name]
                existing_data = []
                row_link_map = {}
                header_row = None
                time_col_idx = None
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if i == 1:
                        header_row = list(row)
                        if '链接' not in header_row or '发布时间' not in header_row:
                            print(f"  警告：工作表 '{name}' 缺少必要的列，跳过更新。")
                            break
                        time_col_idx = header_row.index('发布时间') + 1
                        continue
                    if len(row) == 0:
                        continue
                    row_dict = dict(zip(header_row, row))
                    link = row_dict.get('链接')
                    if link:
                        row_link_map[link] = i
                        existing_data.append(row_dict)

                if time_col_idx is None:
                    continue

                update_count = 0
                append_count = 0

                for _, new_row in df.iterrows():
                    new_link = new_row.get('链接')
                    if not new_link:
                        continue
                    if new_link in row_link_map:
                        row_num = row_link_map[new_link]
                        old_row_dict = next((item for item in existing_data if item.get('链接') == new_link), None)
                        if old_row_dict is None:
                            continue
                        old_time = old_row_dict.get('发布时间', '')
                        new_time = new_row.get('发布时间', '')
                        if is_fuzzy_time(old_time):
                            ws.cell(row=row_num, column=time_col_idx, value=new_time)
                            update_count += 1
                    else:
                        start_row = ws.max_row + 1
                        new_row_dict = new_row.to_dict()
                        row_values = [new_row_dict.get(col, '') for col in header_row]
                        for col_idx, val in enumerate(row_values, start=1):
                            ws.cell(row=start_row, column=col_idx, value=val)
                        append_count += 1

                if update_count > 0:
                    print(f"  工作表 '{name}' 更新了 {update_count} 条记录的发布时间（模糊→精确/更新）。")
                if append_count > 0:
                    print(f"  工作表 '{name}' 新增了 {append_count} 条记录。")
                if update_count == 0 and append_count == 0:
                    print(f"  工作表 '{name}' 无变化。")
            else:
                ws = wb.create_sheet(title=name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                print(f"  已新建工作表 '{name}'，写入 {len(df)} 条记录。")
        wb.save(full)
        print(f"文件已保存: {full}")


# ==================== 主程序 ====================
if __name__ == "__main__":
    # 在程序启动时加载一次 Cookie
    print("=" * 40)
    print("乙方宝采购公告API爬虫程序")
    print("By: Hunter Chen")
    print("在任何输入提示下输入 @restart 可重新开始")
    print("=" * 40)
    print("\n正在检查登录状态...")
    cj = load_cookies()
    if global_driver:
        print("注意：浏览器已保持打开，您可以在其中新建标签页手工验证公告。")

    while True:  # 主循环，支持重启
        try:
            # 如果之前有浏览器实例，确保关闭（重启时）
            # 注意：重启时不关闭浏览器，保持打开状态
            # if global_driver:
            #     try:
            #         global_driver.quit()
            #     except:
            #         pass
            #     global_driver = None

            # --- 1. 运行模式选择 ---
            print("\n请选择运行模式：")
            print("1. 互动模式（运行结束后需按回车退出）")
            print("2. 静默模式（运行结束后自动退出，并可选择自动关机）")
            while True:
                mode_choice = input_with_restart("请输入选项 (1/2，默认1): ").strip()
                if mode_choice == '':
                    mode_choice = '1'
                if mode_choice in ['1', '2']:
                    break
                print("输入无效，请输入1或2。")
            interactive_mode = (mode_choice == '1')
            silent_mode = (mode_choice == '2')

            auto_shutdown = False
            if silent_mode:
                print("\n是否在爬取完成后自动关机？")
                while True:
                    shutdown_choice = input_with_restart("请输入选项 (1: 不关机, 2: 自动关机，默认1): ").strip()
                    if shutdown_choice == '':
                        shutdown_choice = '1'
                    if shutdown_choice in ['1', '2']:
                        break
                    print("输入无效，请输入1或2。")
                auto_shutdown = (shutdown_choice == '2')
                if auto_shutdown:
                    print("将在爬取完成后60秒自动关机。")
                else:
                    print("爬取完成后不会自动关机。")

            # --- 2. 内容类型选择 ---
            print("\n请选择要搜索的内容类型（可多选，输入数字并用逗号或空格分隔）：")
            print("1. 招标信息")
            print("2. 中标信息")
            print("3. 采购意向")
            while True:
                type_input = input_with_restart("请输入选项 (例如：1,2 或 1 2 3，默认1 2): ").strip()
                if type_input == '':
                    type_input = '1 2'
                selected_nums = re.findall(r'\d+', type_input)
                selected_displays = [CONTENT_TYPE_DISPLAY_MAP.get(num) for num in selected_nums if
                                     num in CONTENT_TYPE_DISPLAY_MAP]
                if selected_displays:
                    break
                print("输入无效，请至少选择一个有效选项。")
            print(f"将爬取: {', '.join(selected_displays)}")

            # --- 3. 采购意向的预采时间选择（如果包含采购意向）---
            purchase_params = {}
            if '采购意向' in selected_displays:
                print("\n请选择采购意向的预采时间范围：")
                print("1. 不限")
                print("2. 本月")
                print("3. 下月")
                print("4. 自定义（输入起始年月和结束年月）")
                while True:
                    pt_choice = input_with_restart("请输入选项 (1/2/3/4，默认1): ").strip()
                    if pt_choice == '':
                        pt_choice = '1'
                    if pt_choice in ['1', '2', '3', '4']:
                        break
                    print("输入无效，请输入1-4之间的数字。")
                if pt_choice == '1':
                    purchase_params['采购意向'] = {'type': '-1', 'times': None}
                elif pt_choice == '2':
                    purchase_params['采购意向'] = {'type': '1', 'times': None}
                elif pt_choice == '3':
                    purchase_params['采购意向'] = {'type': '2', 'times': None}
                else:  # 自定义
                    current_year, current_month = get_current_year_month()
                    max_end_year, max_end_month = get_three_months_later()
                    while True:
                        start_ym = input_with_restart("请输入起始年月（例如 2026-01 或 2026年1月）: ").strip()
                        if not start_ym:
                            print("起始年月不能为空。")
                            continue
                        parsed_start = parse_year_month(start_ym)
                        if not parsed_start:
                            print("格式无效，请使用 YYYY-MM 或 YYYY年M月格式。")
                            continue
                        start_year, start_month = map(int, parsed_start.split('/'))
                        if (start_year, start_month) > (current_year, current_month):
                            print(f"起始年月不能晚于当前月份 {current_year}年{current_month}月，请重新输入。")
                            continue
                        break
                    while True:
                        end_ym = input_with_restart("请输入结束年月（例如 2026-03 或 2026年3月）: ").strip()
                        if not end_ym:
                            print("结束年月不能为空。")
                            continue
                        parsed_end = parse_year_month(end_ym)
                        if not parsed_end:
                            print("格式无效，请使用 YYYY-MM 或 YYYY年M月格式。")
                            continue
                        end_year, end_month = map(int, parsed_end.split('/'))
                        if (end_year, end_month) > (max_end_year, max_end_month):
                            print(f"结束年月不能晚于 {max_end_year}年{max_end_month}月（当前月份后3个月），请重新输入。")
                            continue
                        if (end_year, end_month) < (start_year, start_month):
                            print("结束年月不能早于起始年月，请重新输入。")
                            continue
                        break
                    purchase_params['采购意向'] = {'type': '', 'times': f"{parsed_start},{parsed_end}"}

            # --- 4. 地区选择 ---
            print("\n请选择地区：")
            print("1. 全国")
            print("2. 指定省份（可多个，请用中文分号；分隔，例如：湖南；广东）")
            print("3. 南区（广东；广西；海南；湖南；江西）")
            while True:
                region_choice = input_with_restart("请输入选项 (1/2/3，默认3): ").strip()
                if region_choice == '':
                    region_choice = '3'
                if region_choice in ['1', '2', '3']:
                    break
                print("输入无效，请输入1、2或3。")

            province_ids = ""
            province_names = []
            if region_choice == '1':
                province_names = ["全国"]
                province_ids = PROVINCE_ID_MAP["全国"]
            elif region_choice == '3':
                province_names = SOUTH_PROVINCES
                province_ids = ",".join([PROVINCE_ID_MAP[p] for p in SOUTH_PROVINCES])
                print(f"已选择南区省份：{', '.join(province_names)}")
            else:
                while True:
                    region_input = input_with_restart("请输入省份名称（多个用中文分号；分隔，如：湖南；广东）: ").strip()
                    if region_input:
                        names = [r.strip() for r in re.split(r'[；;]', region_input) if r.strip()]
                        valid_ids = []
                        valid_names = []
                        for name in names:
                            if name in PROVINCE_ID_MAP:
                                valid_ids.append(PROVINCE_ID_MAP[name])
                                valid_names.append(name)
                            else:
                                print(f"警告：未找到省份 '{name}' 的ID，已忽略。")
                        if valid_ids:
                            province_ids = ",".join(valid_ids)
                            province_names = valid_names
                            break
                        else:
                            print("未输入有效省份，请重新输入。")
                    else:
                        print("省份名称不能为空，请重新输入。")
                print(f"将使用地区ID组合: {province_ids} (对应: {', '.join(province_names)})")

            # --- 5. 检索位置（默认全文检索）---
            print("\n请选择检索位置：")
            print("1. 全文检索")
            print("2. 标题检索")
            pos_choice = input_with_restart("请输入选项 (1/2，直接回车默认为1): ").strip()
            if pos_choice == '':
                pos_choice = '1'
            search_position = "全文检索" if pos_choice == '1' else "标题检索"
            print(f"已选择: {search_position}")

            # --- 6. 搜索模式 ---
            print("\n请选择搜索模式：")
            print("1. 使用默认关键词组合")
            print("2. 自定义输入关键词（多个关键词请用逗号分隔，例如：DSA,核磁,CT）")
            while True:
                choice = input_with_restart("请输入选项 (1/2，默认1): ").strip()
                if choice == '':
                    choice = '1'
                if choice in ['1', '2']:
                    break
                print("输入无效，请输入1或2。")
            search_terms = []
            if choice == '2':
                while True:
                    custom_kw = input_with_restart(
                        "请输入自定义关键词（多个用中文或英文逗号分隔，例如：DSA,核磁,CT）: ").strip()
                    if custom_kw:
                        search_terms = [term.strip() for term in re.split(r'[，,]', custom_kw) if term.strip()]
                        break
                    else:
                        print("关键词不能为空，请重新输入。")
                print(f"将使用自定义关键词: {search_terms}")
            else:
                all_default_kw = KEYWORDS_DSA + KEYWORDS_CT + KEYWORDS_MR + KEYWORDS_OTHER
                search_terms = list(dict.fromkeys(all_default_kw))
                print(f"将使用默认关键词: {search_terms}")

            # --- 7. 定时监控询问 ---
            monitor_mode = False
            monitor_interval = 0
            if interactive_mode:
                print("\n是否启用定时监控模式？")
                while True:
                    monitor_choice = input_with_restart("请输入选项 (1: 仅单次抓取, 2: 定时监控，默认1): ").strip()
                    if monitor_choice == '':
                        monitor_choice = '1'
                    if monitor_choice in ['1', '2']:
                        break
                    print("输入无效，请输入1或2。")
                if monitor_choice == '2':
                    while True:
                        interval_input = input_with_restart("请输入监控周期（分钟，输入0表示无间隔连续抓取）: ").strip()
                        if interval_input.isdigit() and int(interval_input) >= 0:
                            monitor_interval = int(interval_input)
                            monitor_mode = True
                            if monitor_interval == 0:
                                print("监控间隔为0，将连续抓取（无等待）。")
                            else:
                                print(f"已启用定时监控，周期为 {monitor_interval} 分钟。")
                            break
                        else:
                            print("输入无效，请输入非负整数。")

            # 监控模式下的显示模式选择
            if monitor_mode:
                print("\n请选择监控结果显示模式：")
                print("1. 显示全部结果")
                print("2. 仅显示CT类、DSA类、MR类结果")
                while True:
                    display_mode_choice = input_with_restart("请输入选项 (1/2，默认1): ").strip()
                    if display_mode_choice == '':
                        display_mode_choice = '1'
                    if display_mode_choice in ['1', '2']:
                        break
                    print("输入无效，请输入1或2。")
                display_mode = int(display_mode_choice)
            else:
                display_mode = 1

            # --- 8. 发布日期范围 ---
            start_dt = end_dt = None
            time_option = None
            types_without_caigou = [d for d in selected_displays if d != '采购意向']
            if types_without_caigou:
                print("\n请选择招标/中标信息的发布日期范围：")
                print("1. 当日")
                print("2. 近7天")
                print("3. 近一个月")
                print("4. 近三个月")
                print("5. 近六个月")
                print("6. 自定义日期范围")
                while True:
                    date_option = input_with_restart("请输入选项 (1/2/3/4/5/6，默认1): ").strip()
                    if date_option == '':
                        date_option = '1'
                    if date_option in ['1', '2', '3', '4', '5', '6']:
                        break
                    print("输入无效，请输入1-6之间的数字。")

                if date_option == '1':
                    start_dt, end_dt = get_date_range('1')
                    print(f"将搜索 {start_dt} 至 {end_dt} 范围内的公告")
                elif date_option == '6':
                    print("请输入自定义日期范围（格式如 2026-03-01 或 2026/03/01）：")
                    while True:
                        start_input = input_with_restart("开始日期: ").strip()
                        if not start_input:
                            print("开始日期不能为空，请重新输入。")
                            continue
                        parsed_start = parse_date_input(start_input)
                        if parsed_start:
                            start_dt = parsed_start
                            break
                        else:
                            print("日期格式无效，请使用 YYYY-MM-DD 或 YYYY/MM/DD 格式。")
                    while True:
                        end_input = input_with_restart("结束日期: ").strip()
                        if not end_input:
                            print("结束日期不能为空，请重新输入。")
                            continue
                        parsed_end = parse_date_input(end_input)
                        if parsed_end:
                            end_dt = parsed_end
                            break
                        else:
                            print("日期格式无效，请使用 YYYY-MM-DD 或 YYYY/MM/DD 格式。")
                    print(f"将搜索 {start_dt} 至 {end_dt} 范围内的公告")
                elif date_option in TIME_OPTION_MAP:
                    time_option = TIME_OPTION_MAP[date_option]
                    print("将使用预设日期范围")
                else:
                    date_option = '2'
                    time_option = TIME_OPTION_MAP['2']
                    print("将使用近7天作为日期范围。")
            else:
                print("您只选择了采购意向，无需输入发布日期范围。")

            # --- 9. 是否抓取详情 ---
            print("\n是否抓取公告的详细信息（如医院名称、开标时间、标的详情等）？")
            print("注意：这会大幅增加运行时间")
            while True:
                detail_choice = input_with_restart("请输入选项 (1: 仅列表页, 2: 抓取详情页，默认2): ").strip()
                if detail_choice == '':
                    detail_choice = '2'
                if detail_choice in ['1', '2']:
                    break
                print("输入无效，请输入1或2。")
            grab_detail = (detail_choice == '2')

            # --- 10. 输出文件模式 ---
            print("\n请选择输出文件模式：")
            print("1. 新建文件（文件名自动加时间戳）")
            print("2. 合并到已有文件")
            while True:
                file_mode = input_with_restart("请输入选项 (1/2，默认1): ").strip()
                if file_mode == '':
                    file_mode = '1'
                if file_mode in ['1', '2']:
                    break
                print("输入无效，请输入1或2。")

            output_filename = EXCEL_FILE_NAME
            if file_mode == '2':
                all_files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx')]
                temp_to_main = {}
                for f in all_files:
                    if f.startswith('~$'):
                        main_name = f[2:]
                        temp_to_main[f] = main_name
                print("找到以下Excel文件：")
                display_list = []
                for f in all_files:
                    if f.startswith('~$'):
                        display_list.append(f"临时文件 -> {f} (对应主文件: {temp_to_main[f]})")
                    else:
                        display_list.append(f)
                for i, f in enumerate(display_list, 1):
                    print(f"  {i}. {f}")

                while True:
                    choice = input_with_restart(
                        "请选择要合并到的文件序号（输入数字，或输入 q 放弃合并并新建文件）: ").strip()
                    if choice.lower() == 'q':
                        print("放弃合并，将新建文件。")
                        output_filename = EXCEL_FILE_NAME
                        break
                    if choice.isdigit():
                        idx = int(choice) - 1
                        if 0 <= idx < len(all_files):
                            selected_file = all_files[idx]
                            if selected_file.startswith('~$'):
                                print(
                                    f"您选择了临时文件 {selected_file}，请关闭对应的主文件 {temp_to_main[selected_file]} 后重新选择主文件。")
                                continue
                            temp_file = f"~${selected_file}"
                            if temp_file in all_files:
                                full_path = os.path.join(OUTPUT_DIR, selected_file)
                                try:
                                    with open(full_path, 'ab') as f:
                                        pass
                                    output_filename = selected_file
                                    print(f"已选择文件：{selected_file}")
                                    break
                                except PermissionError:
                                    print(
                                        f"文件 {selected_file} 可能已被其他程序打开（检测到临时文件），请关闭后按回车重试，或输入 q 放弃。")
                                    input("按回车重试...")
                                    continue
                                except Exception as e:
                                    print(f"无法访问文件 {selected_file}：{e}，请选择其他文件或输入 q 放弃。")
                                    continue
                            else:
                                output_filename = selected_file
                                print(f"已选择文件：{selected_file}")
                                break
                        else:
                            print("序号无效，请重新输入。")
                    else:
                        print("输入无效，请输入数字序号或 q。")
                if not output_filename:
                    output_filename = EXCEL_FILE_NAME
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                output_filename = f"{EXCEL_FILE_NAME.rsplit('.', 1)[0]}_{timestamp}.xlsx"
                print(f"将新建文件：{output_filename}")

            # --- 开始爬取 ---
            print("\n开始执行爬虫任务...")


            # 定义爬取函数（内部使用全局 cj）
            def perform_crawl():
                all_type_data = {d: pd.DataFrame() for d in selected_displays}
                for display in selected_displays:
                    ctype = CONTENT_TYPE_API_MAP[display]
                    print(f"\n========== 正在爬取类型: {display} ==========")

                    base_cols = ['关键词', '标题', '公告类型', '发布时间', '城市', '省份', '内容类型', '链接']
                    if display == '招标信息':
                        detail_cols = ['纯公告链接', '医院名称', '开标时间', '采购编号', '标的名称', '数量', '预算金额',
                                       '最高限价']
                    elif display == '中标信息':
                        detail_cols = ['纯公告链接', '医院名称', '采购编号', '标的名称', '品牌', '规格型号', '数量',
                                       '单价', '预算金额', '中标金额', '供应商名称']
                    else:  # 采购意向
                        detail_cols = ['医院名称', '采购项目名称', '采购品目', '采购需求概况', '预算金额',
                                       '预计采购日期']
                    cols = base_cols + detail_cols + ['公告详情', '附件']

                    type_df = pd.DataFrame()

                    for kw in search_terms:
                        print(f"\n===== 正在处理关键词: '{kw}' =====")
                        page = 1
                        kw_data = []
                        while True:
                            if display == '采购意向':
                                pur = purchase_params.get('采购意向', {'type': '-1', 'times': None})
                                url = build_api_url(
                                    province_ids=province_ids,
                                    keyword=kw,
                                    content_type=ctype,
                                    search_type=search_position,
                                    time_option=None,
                                    page_num=page,
                                    purchase_time_type=pur['type'],
                                    purchase_times=pur['times']
                                )
                            else:
                                url = build_api_url(
                                    province_ids=province_ids,
                                    keyword=kw,
                                    content_type=ctype,
                                    search_type=search_position,
                                    time_option=time_option if time_option else "",
                                    page_num=page,
                                    start_date=start_dt,
                                    end_date=end_dt
                                )
                            print(f"  请求第{page}页: {url}")
                            data = fetch_search_results(url)
                            if not data or not data.get("resultList"):
                                print("  无更多数据或请求失败")
                                break
                            items = data["resultList"]
                            print(f"  本页获取到 {len(items)} 条记录")
                            for item in items:
                                raw_time = item.get('updateDate') or item.get('releaseTime') or ''
                                pub_time = convert_fuzzy_time(raw_time)

                                info = {
                                    '关键词': kw,
                                    '标题': re.sub(r'<[^>]+>', '', item.get('title', '') or ''),
                                    '公告类型': item.get('type', '') or '',
                                    '发布时间': pub_time,
                                    '城市': '',
                                    '省份': '',
                                    '内容类型': '',
                                    '链接': '',
                                    '纯公告链接': '', '医院名称': '', '开标时间': '', '采购编号': '',
                                    '标的名称': '', '数量': '', '预算金额': '', '最高限价': '',
                                    '品牌': '', '规格型号': '', '单价': '', '中标金额': '', '供应商名称': '',
                                    '公告详情': '', '附件': '',
                                    '采购项目名称': '', '采购品目': '', '采购需求概况': '', '预计采购日期': '',
                                }

                                if display == '采购意向':
                                    zhao_biao_unit = item.get('zhaoBiaoUnit') or item.get('zhaoBiaoUnitComplate')
                                    info['医院名称'] = re.sub(r'<[^>]+>', '', zhao_biao_unit or '')
                                    budget_amount = item.get('budgetAmount')
                                    if budget_amount:
                                        info['预算金额'] = budget_amount
                                    else:
                                        budget_val = item.get('budget')
                                        if budget_val is not None:
                                            info['预算金额'] = str(budget_val) + (
                                                "万元" if isinstance(budget_val, (int, float)) else "")
                                    info['预计采购日期'] = item.get('estimateTime', '') or ''
                                    purchase_survey = item.get('purchaseSurvey')
                                    if purchase_survey is not None:
                                        info['采购需求概况'] = re.sub(r'<[^>]+>', '', purchase_survey)
                                    info['采购项目名称'] = info['标题']
                                    rel_id = item.get('relationId')
                                    a_id = item.get('areaId')
                                    if rel_id and a_id:
                                        info[
                                            '链接'] = f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{rel_id}/{a_id}"
                                    area_id = str(item.get('areaId', ''))
                                    info['省份'] = ID_TO_PROVINCE.get(area_id, '')
                                    area_str = item.get('area', '')
                                    if area_str:
                                        parts = area_str.split('-')
                                        if len(parts) >= 2:
                                            info['城市'] = parts[1].strip()
                                        else:
                                            info['城市'] = parts[0].strip()
                                else:
                                    area_name = item.get('areaName', '')
                                    if area_name:
                                        parts = area_name.split('-')
                                        if len(parts) >= 2:
                                            info['省份'] = parts[0].strip()
                                            info['城市'] = parts[1].strip()
                                        else:
                                            info['省份'] = parts[0].strip()
                                    if not info['省份']:
                                        area_id = str(item.get('areaId', ''))
                                        info['省份'] = ID_TO_PROVINCE.get(area_id, '')
                                    cid = item.get('contentId')
                                    aid = item.get('areaId')
                                    if cid and aid:
                                        info[
                                            '链接'] = f"https://qiye.qianlima.com/new_qd_yfbsite/#/infoCenter/infoDetail/{cid}/{aid}"

                                if grab_detail and info['链接']:
                                    detail = extract_detail_info(info['链接'], ctype)
                                    if detail:
                                        info.update(detail)

                                # 根据标题和详情判断内容类型
                                detail_text = info.get('公告详情', '')
                                info['内容类型'] = determine_content_type(info.get('标题', ''), detail_text)

                                kw_data.append(info)

                            total = data.get('count', 0)
                            if len(kw_data) >= total or len(items) < 30:
                                print("  已到达最后一页")
                                break
                            page += 1
                            time.sleep(random.uniform(1, 2))

                        if kw_data:
                            kw_df = pd.DataFrame(kw_data)
                            type_df = pd.concat([type_df, kw_df], ignore_index=True)
                            print(f"关键词 '{kw}' 抓取完成，当前类型累计 {len(type_df)} 条")
                        else:
                            print(f"关键词 '{kw}' 未抓取到数据")

                    if not type_df.empty:
                        before = len(type_df)
                        type_df.drop_duplicates(subset=['链接'], keep='first', inplace=True)
                        print(f"  类型 '{display}' 内部去重完成，移除了 {before - len(type_df)} 条重复记录。")
                        for col in cols:
                            if col not in type_df.columns:
                                type_df[col] = ''
                        type_df = type_df[cols]
                        all_type_data[display] = type_df
                        print(f"类型 '{display}' 爬取完成，共 {len(type_df)} 条记录")
                    else:
                        print(f"类型 '{display}' 未获取到数据")
                return all_type_data


            # 执行首次爬取
            all_type_data = perform_crawl()
            all_data_combined = pd.concat([df for df in all_type_data.values() if not df.empty], ignore_index=True)

            if monitor_mode:
                non_empty = {n: df for n, df in all_type_data.items() if not df.empty}
                if non_empty:
                    if file_mode == '1':
                        save_multi_sheet_to_excel(non_empty, output_filename, mode='new')
                    else:
                        full_path = os.path.join(OUTPUT_DIR, output_filename)
                        if os.path.exists(full_path):
                            while True:
                                try:
                                    with open(full_path, 'ab') as f:
                                        pass
                                    break
                                except PermissionError:
                                    print(f"文件 {output_filename} 已被其他程序打开，请关闭后按回车继续...")
                                    input("按回车重试...")
                                except Exception as e:
                                    print(f"无法访问文件 {output_filename}：{e}")
                                    break
                        save_multi_sheet_to_excel(non_empty, output_filename, mode='merge')

                set_console_window_top()
                print("\n" + "=" * 40)
                print("首次抓取完成，结果清单：")
                if display_mode == 1:
                    for display, df in all_type_data.items():
                        if df.empty:
                            continue
                        print(f"\n【{display}】共 {len(df)} 条公告")
                        for idx, row in df.iterrows():
                            title = row.get('标题', '')
                            link = row.get('链接', '')
                            print(f"  {idx + 1}. {title}")
                            print(f"     链接：{link}")
                else:
                    has_device = False
                    exclude_classes = ['工程类', '服务类', '仪器类']
                    device_classes = ['CT类', 'DSA类', 'MR类']
                    device_keywords_lower = [kw.lower() for kw in KEYWORDS_CT + KEYWORDS_DSA + KEYWORDS_MR]
                    for display, df in all_type_data.items():
                        if df.empty:
                            continue
                        device_rows = []
                        for _, row in df.iterrows():
                            content_type_val = row.get('内容类型', '')
                            if any(cls in content_type_val for cls in exclude_classes):
                                continue
                            if any(cls in content_type_val for cls in device_classes):
                                device_rows.append(row)
                                continue
                            detail_text = row.get('公告详情', '')
                            if detail_text:
                                detail_lower = detail_text.lower()
                                if any(kw in detail_lower for kw in device_keywords_lower):
                                    device_rows.append(row)
                        if device_rows:
                            has_device = True
                            print(f"\n【{display}】共 {len(device_rows)} 条设备类公告")
                            for idx, row in enumerate(device_rows, 1):
                                title = row.get('标题', '')
                                link = row.get('链接', '')
                                print(f"  {idx}. {title}")
                                print(f"     链接：{link}")
                    if not has_device:
                        print("（本次未抓取到CT/DSA/MR相关的公告）")
                print("=" * 40)

                try:
                    input_with_restart("\n请查看以上公告，复制链接后按回车键继续...")
                except RestartException:
                    # 重启时直接回到主循环开头
                    continue

                print("\n是否进入监控模式，开始周期性抓取？")
                while True:
                    cont = input_with_restart("请输入选项 (1: 停止监控, 2: 开始监控，默认2): ").strip()
                    if cont == '':
                        cont = '2'
                    if cont in ['1', '2']:
                        break
                    print("输入无效，请输入1或2。")
                if cont == '2':
                    previous_links = set(
                        all_data_combined['链接'].dropna().unique()) if not all_data_combined.empty else set()
                    total_all_data = all_type_data

                    print("\n进入监控模式...")
                    while True:
                        print("开始新一轮监控抓取...")
                        new_all_type_data = perform_crawl()
                        new_combined = pd.concat([df for df in new_all_type_data.values() if not df.empty],
                                                 ignore_index=True)
                        new_links = set(new_combined['链接'].dropna().unique()) if not new_combined.empty else set()

                        added_links = new_links - previous_links

                        if added_links:
                            for display in total_all_data:
                                if display in new_all_type_data:
                                    total_all_data[display] = pd.concat(
                                        [total_all_data[display], new_all_type_data[display]], ignore_index=True)
                                    total_all_data[display].drop_duplicates(subset=['链接'], keep='first', inplace=True)

                            non_empty = {n: df for n, df in total_all_data.items() if not df.empty}
                            if non_empty:
                                save_multi_sheet_to_excel(non_empty, output_filename, mode='merge')

                            set_console_window_top()
                            print("\n" + "!" * 50)
                            print(f"检测到 {len(added_links)} 条新公告！")
                            print("!" * 50)

                            if display_mode == 1:
                                new_records_list = []
                                for display in new_all_type_data:
                                    df = new_all_type_data[display]
                                    if df.empty:
                                        continue
                                    new_in_df = df[df['链接'].isin(added_links)]
                                    for _, row in new_in_df.iterrows():
                                        new_records_list.append({
                                            '类型': display,
                                            '标题': row.get('标题', ''),
                                            '链接': row.get('链接', '')
                                        })
                                if new_records_list:
                                    print("新增公告详情：")
                                    for i, rec in enumerate(new_records_list, 1):
                                        print(f"  {i}. [{rec['类型']}] {rec['标题']}")
                                        print(f"     链接：{rec['链接']}")
                                else:
                                    print("（无法获取新增记录的详情）")

                                try:
                                    input_with_restart("\n请查看以上公告，复制链接后按回车键继续...")
                                except RestartException:
                                    # 重启时跳出监控循环，回到主循环开头
                                    break

                                while True:
                                    cont = input_with_restart("\n是否继续监控？(1: 停止, 2: 继续，默认2): ").strip()
                                    if cont == '':
                                        cont = '2'
                                    if cont in ['1', '2']:
                                        break
                                    print("输入无效，请输入1或2。")
                                if cont == '1':
                                    break
                                else:
                                    previous_links = new_links
                                    if monitor_interval > 0:
                                        for remaining in range(monitor_interval * 60, 0, -1):
                                            mins, secs = divmod(remaining, 60)
                                            timer = f'{mins:02d}:{secs:02d}'
                                            print(f'\r下一轮监控倒计时: {timer} ', end='', flush=True)
                                            time.sleep(1)
                                        print()
                                    continue
                            else:
                                new_device_records = []
                                exclude_classes = ['工程类', '服务类', '仪器类']
                                device_classes = ['CT类', 'DSA类', 'MR类']
                                device_keywords_lower = [kw.lower() for kw in KEYWORDS_CT + KEYWORDS_DSA + KEYWORDS_MR]
                                for display in new_all_type_data:
                                    df = new_all_type_data[display]
                                    if df.empty:
                                        continue
                                    new_in_df = df[df['链接'].isin(added_links)]
                                    for _, row in new_in_df.iterrows():
                                        content_type_val = row.get('内容类型', '')
                                        if any(cls in content_type_val for cls in exclude_classes):
                                            continue
                                        if any(cls in content_type_val for cls in device_classes):
                                            new_device_records.append({
                                                '类型': display,
                                                '标题': row.get('标题', ''),
                                                '链接': row.get('链接', '')
                                            })
                                            continue
                                        detail_text = row.get('公告详情', '')
                                        if detail_text:
                                            detail_lower = detail_text.lower()
                                            if any(kw in detail_lower for kw in device_keywords_lower):
                                                new_device_records.append({
                                                    '类型': display,
                                                    '标题': row.get('标题', ''),
                                                    '链接': row.get('链接', '')
                                                })

                                if new_device_records:
                                    print("新增设备类公告详情：")
                                    for i, rec in enumerate(new_device_records, 1):
                                        print(f"  {i}. [{rec['类型']}] {rec['标题']}")
                                        print(f"     链接：{rec['链接']}")

                                    try:
                                        input_with_restart("\n请查看以上设备类公告，复制链接后按回车键继续...")
                                    except RestartException:
                                        break

                                    while True:
                                        cont = input_with_restart("\n是否继续监控？(1: 停止, 2: 继续，默认2): ").strip()
                                        if cont == '':
                                            cont = '2'
                                        if cont in ['1', '2']:
                                            break
                                        print("输入无效，请输入1或2。")
                                    if cont == '1':
                                        break
                                    else:
                                        previous_links = new_links
                                        if monitor_interval > 0:
                                            for remaining in range(monitor_interval * 60, 0, -1):
                                                mins, secs = divmod(remaining, 60)
                                                timer = f'{mins:02d}:{secs:02d}'
                                                print(f'\r下一轮监控倒计时: {timer} ', end='', flush=True)
                                                time.sleep(1)
                                            print()
                                        continue
                                else:
                                    print("\n=====本轮新增公告中未包含CT/DSA/MR相关条目 =====")
                                    previous_links = new_links
                                    if monitor_interval > 0:
                                        for remaining in range(monitor_interval * 60, 0, -1):
                                            mins, secs = divmod(remaining, 60)
                                            timer = f'{mins:02d}:{secs:02d}'
                                            print(f'\r下一轮监控倒计时: {timer} ', end='', flush=True)
                                            time.sleep(1)
                                        print()
                                    continue
                        else:
                            print(f"\n=====本轮未发现新公告 =====")
                            if monitor_interval == 0:
                                print("监控间隔为0，立即开始下一轮...")
                                previous_links = new_links
                                continue
                            else:
                                for remaining in range(monitor_interval * 60, 0, -1):
                                    mins, secs = divmod(remaining, 60)
                                    timer = f'{mins:02d}:{secs:02d}'
                                    print(f'\r下一轮监控倒计时: {timer} ', end='', flush=True)
                                    time.sleep(1)
                                print()
                                previous_links = new_links
                                continue
                    # 监控结束
                    print("\n监控结束。数据已实时保存至文件。")
                else:
                    print("\n监控未启动。数据已保存。")
            else:
                non_empty = {n: df for n, df in all_type_data.items() if not df.empty}
                if non_empty:
                    if file_mode == '1':
                        save_multi_sheet_to_excel(non_empty, output_filename, mode='new')
                    else:
                        full_path = os.path.join(OUTPUT_DIR, output_filename)
                        if os.path.exists(full_path):
                            while True:
                                try:
                                    with open(full_path, 'ab') as f:
                                        pass
                                    break
                                except PermissionError:
                                    print(f"文件 {output_filename} 已被其他程序打开，请关闭后按回车继续...")
                                    input("按回车重试...")
                                except Exception as e:
                                    print(f"无法访问文件 {output_filename}：{e}")
                                    break
                        save_multi_sheet_to_excel(non_empty, output_filename, mode='merge')
                    total = sum(len(df) for df in non_empty.values())
                    print(f"\n总计抓取 {total} 条记录")
                else:
                    print("本次爬取未获得任何数据")

            print("\n程序运行完毕。")
            if global_driver:
                print("浏览器已保持打开，您可继续手工验证。如需关闭，请手动关闭浏览器窗口。")

            if auto_shutdown:
                print("程序执行完毕，60秒后自动关机...")
                os.system("shutdown /s /f /t 60")
            elif interactive_mode:
                print("\n按回车键退出...")
                try:
                    input_with_restart("按回车键退出...")
                except:
                    pass
                finally:
                    break  # 退出主循环，结束程序

        except RestartException:
            print("\n检测到重启指令 @restart，正在返回主菜单...\n")
            # 不关闭浏览器，保持现有会话
            # if global_driver:
            #     try:
            #         global_driver.quit()
            #     except:
            #         pass
            #     global_driver = None
            continue  # 回到主循环开头